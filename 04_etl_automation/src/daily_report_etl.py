"""
daily_report_etl.py
Automated ETL pipeline that simulates a recurring business process:
  EXTRACT  -> pull raw transaction files (here: multiple daily CSV "drops")
  TRANSFORM -> clean, standardize, and aggregate
  LOAD     -> refresh a formatted Excel report, ready to email/share

This is the kind of script that replaces a manual "download → paste into
Excel → re-format → email" routine — a common RPA/automation use case in
risk & analytics teams.

Run it repeatedly (e.g., via Task Scheduler / cron / Airflow) and it will
pick up any new files dropped into data/incoming/ and produce a fresh report.
"""
import glob
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

BASE = "/home/claude/projects/04_etl_automation"
INCOMING_DIR = f"{BASE}/data/incoming"
PROCESSED_DIR = f"{BASE}/data/processed"
OUTPUT_DIR = f"{BASE}/output"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", bold=True, size=15, color="1F4E78")
BODY_FONT = Font(name="Arial", size=10)


def extract():
    """Pull all daily CSV drops from the incoming folder."""
    files = sorted(glob.glob(f"{INCOMING_DIR}/*.csv"))
    if not files:
        raise FileNotFoundError(f"No files found in {INCOMING_DIR}")
    frames = []
    for f in files:
        df = pd.read_csv(f)
        df["source_file"] = os.path.basename(f)
        frames.append(df)
    combined = pd.concat(frames, ignore_index=True)
    print(f"[EXTRACT] Pulled {len(files)} file(s), {len(combined):,} total rows")
    return combined, files


def transform(df):
    """Clean and standardize the combined data, then compute daily aggregates."""
    df = df.copy()
    df["transaction_date"] = pd.to_datetime(df["transaction_date"])
    df["merchant_category"] = df["merchant_category"].str.strip().str.title()
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce")

    before = len(df)
    df = df.dropna(subset=["amount", "transaction_date"])
    dropped = before - len(df)

    daily_summary = (
        df.groupby(df["transaction_date"].dt.date)
        .agg(
            total_transactions=("amount", "count"),
            total_volume=("amount", "sum"),
            avg_transaction=("amount", "mean"),
            flagged_transactions=("is_flagged", "sum"),
        )
        .reset_index()
        .rename(columns={"transaction_date": "date"})
    )
    daily_summary["total_volume"] = daily_summary["total_volume"].round(2)
    daily_summary["avg_transaction"] = daily_summary["avg_transaction"].round(2)

    category_summary = (
        df.groupby("merchant_category")
        .agg(total_transactions=("amount", "count"), total_volume=("amount", "sum"))
        .reset_index()
        .sort_values("total_volume", ascending=False)
    )
    category_summary["total_volume"] = category_summary["total_volume"].round(2)

    print(f"[TRANSFORM] Cleaned data ({dropped} invalid rows dropped), "
          f"built daily + category summaries")
    return df, daily_summary, category_summary


def load(df, daily_summary, category_summary):
    """Write a formatted Excel report with a chart, ready to distribute."""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Daily Summary"
    ws1["A1"] = "Daily Transaction Summary Report"
    ws1["A1"].font = TITLE_FONT
    ws1["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")

    headers = list(daily_summary.columns)
    for c, h in enumerate(headers, start=1):
        cell = ws1.cell(row=4, column=c, value=h)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for i, row in daily_summary.iterrows():
        for c, val in enumerate(row, start=1):
            ws1.cell(row=5 + i, column=c, value=val).font = BODY_FONT

    for c in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(c)].width = 20

    # Chart: total volume by day
    chart = BarChart()
    chart.title = "Total Transaction Volume by Day"
    last_row = 4 + len(daily_summary)
    data = Reference(ws1, min_col=3, min_row=4, max_row=last_row)
    cats = Reference(ws1, min_col=1, min_row=5, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height, chart.width = 8, 16
    ws1.add_chart(chart, f"A{last_row + 3}")

    ws2 = wb.create_sheet("By Category")
    headers2 = list(category_summary.columns)
    for c, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    for i, row in category_summary.iterrows():
        for c, val in enumerate(row, start=1):
            ws2.cell(row=2 + i, column=c, value=val).font = BODY_FONT
    for c in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 22

    out_path = f"{OUTPUT_DIR}/Daily_Transaction_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(out_path)
    print(f"[LOAD] Report written to {out_path}")
    return out_path


def archive_processed(files):
    """Move processed source files out of 'incoming' so re-runs don't double-count."""
    os.makedirs(PROCESSED_DIR, exist_ok=True)
    for f in files:
        dest = os.path.join(PROCESSED_DIR, os.path.basename(f))
        os.replace(f, dest)
    print(f"[ARCHIVE] Moved {len(files)} file(s) to {PROCESSED_DIR}")


def run_pipeline():
    print(f"\n{'='*60}\nETL run started: {datetime.now().isoformat()}\n{'='*60}")
    raw, files = extract()
    clean, daily_summary, category_summary = transform(raw)
    out_path = load(clean, daily_summary, category_summary)
    archive_processed(files)
    print(f"\nPipeline complete. Report ready at:\n{out_path}\n")
    return out_path


if __name__ == "__main__":
    run_pipeline()
