"""
build_dashboard.py
Builds an Excel credit-risk dashboard from the scored loan portfolio:
 - Raw Data tab (scored loans)
 - Summary tab with SUMIFS/AVERAGEIFS formulas (real formulas, not hardcoded)
 - Charts: default rate by risk grade, exposure by region, vintage trend
"""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = PROJECT_ROOT / "output" / "scored_portfolio.csv"
OUT_PATH = PROJECT_ROOT / "output" / "Credit_Risk_Dashboard.xlsx"
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=16, color="1F4E78")
SUB_FONT = Font(name="Arial", italic=True, size=10, color="666666")
LABEL_FONT = Font(name="Arial", bold=True, size=11)
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autosize(ws, n_cols, min_width=10, max_width=32):
    for c in range(1, n_cols + 1):
        col_letter = get_column_letter(c)
        max_len = min_width
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, max_width)


def main():
    df = pd.read_csv(DATA_PATH)
    wb = Workbook()

    # ---------------- Raw Data tab ----------------
    ws_data = wb.active
    ws_data.title = "Raw Data"
    cols = list(df.columns)
    ws_data.append(cols)
    style_header_row(ws_data, 1, len(cols))
    for _, row in df.iterrows():
        ws_data.append(list(row))
    for r in range(2, ws_data.max_row + 1):
        for c in range(1, len(cols) + 1):
            ws_data.cell(row=r, column=c).font = BODY_FONT
    ws_data.freeze_panes = "A2"
    autosize(ws_data, len(cols))

    n_rows = ws_data.max_row
    grade_col = cols.index("risk_grade") + 1
    default_col = cols.index("is_default") + 1
    amount_col = cols.index("loan_amount") + 1
    region_col = cols.index("region") + 1
    year_col = cols.index("origination_year") + 1
    tier_col = cols.index("predicted_risk_tier") + 1

    grade_letter = get_column_letter(grade_col)
    default_letter = get_column_letter(default_col)
    amount_letter = get_column_letter(amount_col)
    region_letter = get_column_letter(region_col)
    year_letter = get_column_letter(year_col)
    tier_letter = get_column_letter(tier_col)

    data_range_grade = f"'Raw Data'!${grade_letter}$2:${grade_letter}${n_rows}"
    data_range_default = f"'Raw Data'!${default_letter}$2:${default_letter}${n_rows}"
    data_range_amount = f"'Raw Data'!${amount_letter}$2:${amount_letter}${n_rows}"
    data_range_region = f"'Raw Data'!${region_letter}$2:${region_letter}${n_rows}"
    data_range_year = f"'Raw Data'!${year_letter}$2:${year_letter}${n_rows}"
    data_range_tier = f"'Raw Data'!${tier_letter}$2:${tier_letter}${n_rows}"

    # ---------------- Summary tab ----------------
    ws = wb.create_sheet("Portfolio Summary")
    ws["A1"] = "Credit Risk Portfolio Dashboard"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Synthetic consumer loan portfolio — 20,000 loans | Source: generated for demo (see data/loan_portfolio.csv)"
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")

    # --- KPI row ---
    ws["A4"] = "Total Loans"
    ws["B4"] = f"=COUNTA({data_range_default})"
    ws["A5"] = "Total Exposure ($)"
    ws["B5"] = f"=SUM({data_range_amount})"
    ws["A6"] = "Overall Default Rate"
    ws["B6"] = f"=AVERAGE({data_range_default})"
    ws["B6"].number_format = "0.0%"
    ws["B5"].number_format = "$#,##0"
    for r in range(4, 7):
        ws.cell(row=r, column=1).font = LABEL_FONT
        ws.cell(row=r, column=2).font = Font(name="Arial", bold=True, size=12, color="1F4E78")

    # --- Table 1: Default rate & exposure by risk grade ---
    ws["A9"] = "Default Rate & Exposure by Risk Grade"
    ws["A9"].font = LABEL_FONT
    grades = sorted(df["risk_grade"].dropna().unique().tolist())
    headers1 = ["Risk Grade", "Loan Count", "Total Exposure ($)", "Default Rate"]
    for c, h in enumerate(headers1, start=1):
        ws.cell(row=10, column=c, value=h)
    style_header_row(ws, 10, len(headers1))
    for i, grade in enumerate(grades):
        r = 11 + i
        ws.cell(row=r, column=1, value=grade).font = BODY_FONT
        ws.cell(row=r, column=2, value=f'=COUNTIF({data_range_grade},A{r})').font = BODY_FONT
        ws.cell(row=r, column=3, value=f'=SUMIF({data_range_grade},A{r},{data_range_amount})').font = BODY_FONT
        ws.cell(row=r, column=3).number_format = "$#,##0"
        ws.cell(row=r, column=4, value=f'=AVERAGEIF({data_range_grade},A{r},{data_range_default})').font = BODY_FONT
        ws.cell(row=r, column=4).number_format = "0.0%"
    grade_table_end = 10 + len(grades)

    # --- Table 2: Exposure by region ---
    start2 = grade_table_end + 3
    ws.cell(row=start2 - 1, column=1, value="Exposure & Default Rate by Region").font = LABEL_FONT
    regions = sorted(df["region"].dropna().unique().tolist())
    headers2 = ["Region", "Loan Count", "Total Exposure ($)", "Default Rate"]
    for c, h in enumerate(headers2, start=1):
        ws.cell(row=start2, column=c, value=h)
    style_header_row(ws, start2, len(headers2))
    for i, region in enumerate(regions):
        r = start2 + 1 + i
        ws.cell(row=r, column=1, value=region).font = BODY_FONT
        ws.cell(row=r, column=2, value=f'=COUNTIF({data_range_region},A{r})').font = BODY_FONT
        ws.cell(row=r, column=3, value=f'=SUMIF({data_range_region},A{r},{data_range_amount})').font = BODY_FONT
        ws.cell(row=r, column=3).number_format = "$#,##0"
        ws.cell(row=r, column=4, value=f'=AVERAGEIF({data_range_region},A{r},{data_range_default})').font = BODY_FONT
        ws.cell(row=r, column=4).number_format = "0.0%"
    region_table_end = start2 + len(regions)

    # --- Table 3: Vintage trend by origination year ---
    start3 = region_table_end + 3
    ws.cell(row=start3 - 1, column=1, value="Delinquency Trend by Origination Year").font = LABEL_FONT
    years = sorted(df["origination_year"].dropna().unique().tolist())
    headers3 = ["Origination Year", "Loan Count", "Default Rate"]
    for c, h in enumerate(headers3, start=1):
        ws.cell(row=start3, column=c, value=h)
    style_header_row(ws, start3, len(headers3))
    for i, yr in enumerate(years):
        r = start3 + 1 + i
        ws.cell(row=r, column=1, value=int(yr)).font = BODY_FONT
        ws.cell(row=r, column=2, value=f'=COUNTIF({data_range_year},A{r})').font = BODY_FONT
        ws.cell(row=r, column=3, value=f'=AVERAGEIF({data_range_year},A{r},{data_range_default})').font = BODY_FONT
        ws.cell(row=r, column=3).number_format = "0.0%"
    year_table_end = start3 + len(years)

    # --- Table 4: Predicted risk tier distribution (model output) ---
    start4 = year_table_end + 3
    ws.cell(row=start4 - 1, column=1, value="Model-Predicted Risk Tier Distribution").font = LABEL_FONT
    tiers = ["Low", "Moderate", "High", "Severe"]
    headers4 = ["Risk Tier", "Loan Count", "Total Exposure ($)"]
    for c, h in enumerate(headers4, start=1):
        ws.cell(row=start4, column=c, value=h)
    style_header_row(ws, start4, len(headers4))
    for i, tier in enumerate(tiers):
        r = start4 + 1 + i
        ws.cell(row=r, column=1, value=tier).font = BODY_FONT
        ws.cell(row=r, column=2, value=f'=COUNTIF({data_range_tier},A{r})').font = BODY_FONT
        ws.cell(row=r, column=3, value=f'=SUMIF({data_range_tier},A{r},{data_range_amount})').font = BODY_FONT
        ws.cell(row=r, column=3).number_format = "$#,##0"
    tier_table_end = start4 + len(tiers)

    autosize(ws, 6)

    # ---------------- Charts tab ----------------
    ws_chart = wb.create_sheet("Charts")

    # Chart 1: Default rate by risk grade (bar)
    bar1 = BarChart()
    bar1.title = "Default Rate by Risk Grade"
    bar1.y_axis.title = "Default Rate"
    bar1.x_axis.title = "Risk Grade"
    data1 = Reference(ws, min_col=4, min_row=10, max_row=grade_table_end)
    cats1 = Reference(ws, min_col=1, min_row=11, max_row=grade_table_end)
    bar1.add_data(data1, titles_from_data=True)
    bar1.set_categories(cats1)
    bar1.height, bar1.width = 8, 16
    ws_chart.add_chart(bar1, "A1")

    # Chart 2: Exposure by region (pie)
    pie1 = PieChart()
    pie1.title = "Total Exposure by Region"
    data2 = Reference(ws, min_col=3, min_row=start2, max_row=region_table_end)
    cats2 = Reference(ws, min_col=1, min_row=start2 + 1, max_row=region_table_end)
    pie1.add_data(data2, titles_from_data=True)
    pie1.set_categories(cats2)
    pie1.height, pie1.width = 8, 16
    ws_chart.add_chart(pie1, "A18")

    # Chart 3: Vintage default trend (line)
    line1 = LineChart()
    line1.title = "Default Rate Trend by Origination Year"
    line1.y_axis.title = "Default Rate"
    line1.x_axis.title = "Origination Year"
    data3 = Reference(ws, min_col=3, min_row=start3, max_row=year_table_end)
    cats3 = Reference(ws, min_col=1, min_row=start3 + 1, max_row=year_table_end)
    line1.add_data(data3, titles_from_data=True)
    line1.set_categories(cats3)
    line1.height, line1.width = 8, 16
    ws_chart.add_chart(line1, "A35")

    # Chart 4: Risk tier exposure (bar)
    bar2 = BarChart()
    bar2.title = "Exposure at Risk by Model-Predicted Tier"
    bar2.y_axis.title = "Total Exposure ($)"
    bar2.x_axis.title = "Risk Tier"
    data4 = Reference(ws, min_col=3, min_row=start4, max_row=tier_table_end)
    cats4 = Reference(ws, min_col=1, min_row=start4 + 1, max_row=tier_table_end)
    bar2.add_data(data4, titles_from_data=True)
    bar2.set_categories(cats4)
    bar2.height, bar2.width = 8, 16
    ws_chart.add_chart(bar2, "A52")

    wb.save(OUT_PATH)
    print(f"Saved dashboard to {OUT_PATH}")


if __name__ == "__main__":
    main()
